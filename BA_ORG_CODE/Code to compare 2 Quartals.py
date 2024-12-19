import os
import logging
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
from openpyxl import load_workbook
from openpyxl.styles import Font

base_directory = 'C:/Users/A.TANGYIELEMDONFACK/OneDrive - Zurich Insurance/Bachelorarbeit/CodeBA/'
file_old = os.path.join(base_directory, "GI_annuities_data_template.v10_GE_2023_Q1_original.xlsx")
file_new = os.path.join(base_directory, "GI_annuities_data_template.v10_GE_2024_Q1_original.xlsx")
file_paths = [
    os.path.join(base_directory, "GI_annuities_data_template.v10_GE_2023_Q1_original.xlsx"),
    os.path.join(base_directory, "GI_annuities_data_template.v10_GE_2024_Q1_original.xlsx")
]


logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ReserveBerechnung:
    def __init__(self, death_table):
        self.death_table = death_table

    def berechne_reserve(self, alter, rente, uebersterblichkeit, zins, n, sex, escRate, freq, geburtsjahr):
        v = 1 / (1 + (zins / 100))  # Abzinsfaktor

        if sex == 0:  # "0" for male gender
            adjusted_qx = self.death_table["q x+0"] * uebersterblichkeit  # Übersterblichkeit berücksichtigen
            px = [1 - qx for qx in adjusted_qx]
            lx = [1000000]  # Startwert für lx
            for i in range(1, len(px)):
                lx.append(lx[i - 1] * px[i - 1])
            birthyear_row = self.death_table[self.death_table['BIRTH_YEAR'] == geburtsjahr]
            if not birthyear_row.empty:
                altersverschiebung = birthyear_row['AGE_ADJUSTMENT_M'].values[0]
                alterHeute = berechnungsjahr - geburtsjahr
                t_alter = alterHeute + altersverschiebung
            else:
                raise IndexError("Der Geburtsjahr für einen Mann wurde falsch gerechnet oder ist nicht zu finden.")
        elif sex == 1:  # "1" for female gender
            adjusted_qy = self.death_table["q y+0"] * uebersterblichkeit  # Übersterblichkeit berücksichtigen
            py = [1 - qy for qy in adjusted_qy]
            lx = [1000000]  # Startwert für lx (using lx to keep consistency)
            for i in range(1, len(py)):
                lx.append(lx[i - 1] * py[i - 1])
            birthyear_row = self.death_table[self.death_table['BIRTH_YEAR'] == geburtsjahr]
            if not birthyear_row.empty:
                altersverschiebung = birthyear_row['AGE_ADJUSTMENT_F'].values[0]
                alterHeute = berechnungsjahr - geburtsjahr
                t_alter = alterHeute + altersverschiebung
            else:
                raise IndexError("Der Geburtsjahr für eine Frau wurde falsch gerechnet oder ist nicht zu finden.")
        else:
            return "Ungültige Geschlechtseingabe"

        if n == 121:
            n = 121 - t_alter

        if t_alter + n > len(lx):
            raise IndexError("Der Wert von alter + n überschreitet die Länge von lx.")

        if escRate == "YES":
            rente = rente * (1 + (Rate / 100))

        if art == "Nachschüssig":
            barwertfaktor = 0
            barwertfaktornormal = 0
            n = int(n)
            t_alter = int(t_alter)
            for k in range(1, n + 1):
                if t_alter + k >= len(lx):
                    raise IndexError("Index out of bounds while accessing lx.")
                if freq == 1:
                    barwertfaktor += (v ** k) * (lx[t_alter + k] / lx[t_alter])
                else:
                    barwertfaktornormal += (v ** k) * (lx[t_alter + k] / lx[t_alter])
            if freq != 1:
                correction_term = ((((lx[t_alter + n] * (v ** (t_alter + n)))) / (lx[t_alter] * (v ** t_alter))) - 1) * ((freq - 1) / (2 * freq))
                barwertfaktor = barwertfaktornormal - correction_term

            barwert = rente * barwertfaktor
            return barwert
        elif art == "Vorschüssig":
            barwertfaktor = 0
            barwertfaktornormal = 0
            n = int(n)
            t_alter = int(t_alter)
            for k in range(0, n):
                if t_alter + k >= len(lx):
                    raise IndexError("Index out of bounds while accessing lx.")
                if freq == 1:
                    barwertfaktor += (v ** k) * (lx[t_alter + k] / lx[t_alter])
                else:
                    barwertfaktornormal += (v ** k) * (lx[t_alter + k] / lx[t_alter])
            if freq != 1:
                correction_term = ((((lx[t_alter + n] * (v ** (t_alter + n)))) / (lx[t_alter] * (v ** t_alter))) - 1) * ((freq - 1) / (2 * freq))
                barwertfaktor = barwertfaktornormal + correction_term

            barwert = rente * barwertfaktor
            return barwert

    def compare_reserves(self, directory, base_name='GI_annuities_data_template_with_reserves'):
        file_extension = '.xlsx'
        reserve_sums = {}
        row_counts = {} 

        for filename in os.listdir(directory):
            if filename.startswith(base_name) and filename.endswith(file_extension):
                file_path = os.path.join(directory, filename)
                df = pd.read_excel(file_path, sheet_name="Reserves", header=5)

                # Ensure the 'Reserve' column exists
                if 'Reserve' not in df.columns:
                     print(f"File: {filename}")
                     print(f"Columns: {df.columns.tolist()}")
                     raise ValueError(f"Spalte 'Reserve' fehlt in der Datei {filename}.")
                reserve_column_index = df.columns.get_loc('Reserve')
                reserve_sum = df.iloc[-1, reserve_column_index]
                row_count = len(df)
                reserve_sums[filename] = reserve_sum
                row_counts[filename] = row_count
                if 'Reserve' not in df.columns:
                 raise ValueError(f"Spalte 'Reserve' fehlt in der Datei {filename}.")


        comparison_df = pd.DataFrame(list(reserve_sums.items()), columns=['File', 'Sum of Reserves'])
        comparison_df.sort_values(by='File', inplace=True)
        comparison_df['Delta'] = comparison_df['Sum of Reserves'].diff()
        comparison_df['Percentage Change'] = comparison_df['Delta'] / comparison_df['Sum of Reserves'].shift(1) * 100
        comparison_df['Number of Insured'] = comparison_df['File'].map(row_counts)

        comparison_output_path = os.path.join(directory, 'reserves_comparison_with_deltas.xlsx')
        comparison_df.to_excel(comparison_output_path, index=False)
        print(f"Comparison of reserves with deltas saved to: {comparison_output_path}")
        return comparison_df

    def plot_reserve_comparison(self, comparison_df, base_directory):
        plt.figure(figsize=(10, 6))
        plt.plot(comparison_df['File'], comparison_df['Sum of Reserves'], marker='o')
        plt.title('Sum of Reserves Comparison')
        xlabel = f"Interest Rate: {variables_df.at[0, 1]}, Escalation Rate: {variables_df.at[1, 1]}"
        plt.xlabel(xlabel)
        plt.ylabel('Sum of Reserves')
        plt.xticks(rotation=45, ha='right')
        ax = plt.gca()
        ax.yaxis.set_major_formatter(ticker.FuncFormatter(lambda x, _: f'{x:,.2f}'))
        plt.grid(True)
        plt.tight_layout()
        output_path = os.path.join(base_directory, 'reserves_comparison_plot.png')
        plt.savefig(output_path)
        plt.show()
        logger.info(f"Plot saved to: {output_path}")

    def compare_annuity_changes(self, file_old, file_new, output_file):
        df_old= pd.read_excel(file_old, sheet_name="Inputs - MPs", header=5)
        df_new= pd.read_excel(file_new, sheet_name="Inputs - MPs", header=5)

        output_path=os.path.join(output_file,"Bestandsänderung.xlsx")

        with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:

            df_gone_annuities = df_old[~df_old["DAMAGE-ID"].isin(df_new["DAMAGE-ID"])]
            df_gone_annuities.to_excel(writer, sheet_name="Lost Annuities",index=False)
            print("Lost Annuities:")
            print(df_gone_annuities)

            df_new_annuities = df_new[~df_new["DAMAGE-ID"].isin(df_old["DAMAGE-ID"])]
            df_new_annuities.to_excel(writer, sheet_name="New Annuities",index=False)
            print("New Annuities:")
            print(df_new_annuities)


            df_common_ID=df_new[df_new["DAMAGE-ID"].isin(df_old["DAMAGE-ID"])]
            print(df_common_ID)
            
            df_changed=(df_common_ID.merge(df_old,
                                        how="outer", 
                                        indicator=True)
                        .query("_merge=='left_only'")
                        .drop(columns='_merge'))
            df_changed.to_excel(writer, sheet_name="Changed Annuities",index=False)
            print("Changed Annuities:")
            print(df_changed)

        print(f"Die Excel-Datei wurde erstellt: {output_path}")
           

for i, file_path in enumerate(file_paths, start=1):
    bestand_df = pd.read_excel(file_path, sheet_name="Inputs - MPs", header=5)
    death_table_df = pd.read_excel(file_path, sheet_name="Death Table", header=3)
    variables_df = pd.read_excel(file_path, sheet_name="Variables", header=None)

    zins = variables_df.at[0, 1]  # Liest den Wert in Zelle B1 (Zins)
    Rate = variables_df.at[1, 1]  # Liest den Wert in Zelle B2 (EscRate)
    berechnungsjahr = variables_df.at[2, 1]  # Liest den Wert in Zelle B3
    art = variables_df.at[3, 1]  # Liest den Wert in Zelle B4

    reserve_berechnung = ReserveBerechnung(death_table=death_table_df)

    required_columns = ['AGE_AT_ENTRY', 'ANN_ANNUITY', 'POL_TERM_Y', 'SEX', 'ESC_RATE', 'ANNUITY_FREQ', 'ENTRY_YEAR', "Q_CORR_PN"]
    for col in required_columns:
        if col not in bestand_df.columns:
            raise ValueError(f"Missing required column: {col}")

    for index, row in bestand_df.iterrows():
        alter = row["AGE_AT_ENTRY"]
        rente = row["ANN_ANNUITY"]
        uebersterblichkeit = row.get("Q_CORR_PN", 1.0)  # Verwende den Übersterblichkeitsfaktor oder den Standardwert
        n = row["POL_TERM_Y"]  # Verwende die Laufzeit aus der Tabelle
        sex = row["SEX"]
        escRate = row["ESC_RATE"]
        freq = row["ANNUITY_FREQ"]
        geburtsjahr = row["ENTRY_YEAR"] - row["AGE_AT_ENTRY"]

        # Berechnung der Reserve
        barwert = reserve_berechnung.berechne_reserve(alter, rente, uebersterblichkeit, zins, n, sex, escRate, freq, geburtsjahr)
        bestand_df.at[index, 'Reserve'] = barwert

    summe_reserven = bestand_df['Reserve'].sum()

    # Speichern der Ergebnisse in einer neuen Excel-Datei
    base_name = 'GI_annuities_data_template_with_reserves'
    file_extension = '.xlsx'

    new_file_path = f"{base_directory}{base_name}_{i}{file_extension}"
    bestand_df.to_excel(new_file_path, sheet_name="Reserves", index=False, startrow=5)
    
    wb = load_workbook(new_file_path)
    ws = wb['Reserves']

    # Hinzufügen der Summe unterhalb der letzten Zeile der 'Reserve' Spalte
    last_row = len(bestand_df) + 7 
    sum_cell = ws.cell(row=last_row, column=bestand_df.columns.get_loc("Reserve") + 1)
    sum_cell.value = summe_reserven
    sum_cell.font = Font(bold=True)  # Text in Fettformat

    sum_cell_up = ws.cell(row=2, column=bestand_df.columns.get_loc("Reserve") + 1)
    sum_cell_up.value = summe_reserven
    sum_cell_up.font = Font(bold=True)

    ws.cell(row=1, column=2, value=f"Zins: {zins}, Escalation Rate: {Rate}, Jahr: {berechnungsjahr}")
    ws.cell(row=2,column=25,value=f"Summe:")
    
    # Speichern der Datei mit der hinzugefügten Summe
    wb.save(new_file_path)

    print(f"Die Ergebnisse wurden in die neue Datei gespeichert: {new_file_path}")
    logger.info(f"Process complete: {new_file_path}")

output_file = os.path.join(base_directory, "reserves_comparison_with_deltas.xlsx")

# Ausführung der verschiedenen Funktionen
reserve_berechnung = ReserveBerechnung(death_table=death_table_df)
comparison_df = reserve_berechnung.compare_reserves(base_directory)
reserve_berechnung.plot_reserve_comparison(comparison_df, base_directory)
reserve_berechnung.compare_annuity_changes(file_old, file_new, base_directory)

