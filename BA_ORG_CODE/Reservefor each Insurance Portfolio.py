import os
import logging
import pandas as pd
import math
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import Font


logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

 # Klasse für die Reserveberechnung
class ReserveBerechnung:
    def __init__(self, death_table):
        self.death_table = death_table
    
    """
    berechne_reserve: Berechnet die Reserve für eine bestimmte Police.

    Parameter:
    - alter: Alter der versicherten Person
    - rente: Rentenbetrag
    - uebersterblichkeit: Anpassungsfaktor für die Sterblichkeit
    - zins: Zinssatz in Prozent
    - n: Laufzeit der Police in Jahren
    - sex: Geschlecht (0 für männlich, 1 für weiblich)
    - escRate: Eskalationsrate (Flag: 'YES' oder 'NO')
    - freq: Häufigkeit der Rentenzahlung (Annuity frequency)
    - geburtsjahr: Geburtsjahr der versicherten Person

    Rückgabe:
    - barwert: Berechneter Reservebetrag
    """
    
    def berechne_reserve(self, alter, rente, uebersterblichkeit, zins, n, sex, escRate,freq,geburtsjahr):
    
        v = 1 / (1 + (zins/100)) # Abzinsfaktor basierend auf dem Zinssatz
 
        # Geschlechtsspezifische Berechnung der Reserven
        if sex == 0: # Für männliche versicherte Personen
            # Anpassung der Sterbewahrscheinlichkeit anhand der Übersterblichkeit
            adjusted_qx = self.death_table["q x+0"] * uebersterblichkeit 
            px = [1 - qx for qx in adjusted_qx] # Wahrscheinlichkeit, zu überleben
            lx = [1000000] # Startwert für lx, in der Regel 100 Tausend (lx repräsentiert das Überleben)
            for i in range(1, len(px)):
                lx.append(lx[i - 1] * px[i - 1]) # lx-Werte für jedes Jahr basierend auf px

            # Finden der altersbedingten Anpassung aus der Sterbetabelle
            birthyear_row = self.death_table[self.death_table['BIRTH_YEAR'] == geburtsjahr]    
            if not birthyear_row.empty:
                altersverschiebung = birthyear_row['AGE_ADJUSTMENT_M'].values[0]
                alterHeute = berechnungsjahr - geburtsjahr  
                t_alter = alterHeute + altersverschiebung
            else:
                raise IndexError("Der Geburtsjahr für einen Mann wurde falsch gerechnet oder ist nicht zu finden.")
            

        elif sex == 1: # Für weibliche versicherte Personen
            adjusted_qy = self.death_table["q y+0"] * uebersterblichkeit # Übersterblichkeit berücksichtigen
            py = [1 - qy for qy in adjusted_qy] # Wahrscheinlichkeit, zu überleben
            lx = [1000000] # Startwert für lx 
            for i in range(1, len(py)):
                lx.append(lx[i - 1] * py[i - 1]) # lx-Werte für jedes Jahr basierend auf py
            
            # Finden der altersbedingten Anpassung für Frauen
            birthyear_row = self.death_table[self.death_table['BIRTH_YEAR'] == geburtsjahr]    
            if not birthyear_row.empty:
                altersverschiebung = birthyear_row['AGE_ADJUSTMENT_F'].values[0]
                alterHeute = berechnungsjahr - geburtsjahr
                t_alter = alterHeute + altersverschiebung
            else:
                raise IndexError("Der Geburtsjahr für eine Frau wurde falsch gerechnet oder ist nicht zu finden.")
        else:
            return "Ungültige Geschlechtseingabe"
   
        # Überprüfung, ob die Laufzeit korrekt berechnet wird
        if n == 121:
            n = 121 - t_alter

        # Sicherstellen, dass der Wert von alter + n nicht die Länge von lx überschreitet
        if t_alter + n > len(lx):
            raise IndexError("Der Wert von alter + n überschreitet die Länge von lx.")
        
        # Eskalationsrate berücksichtigen, wenn sie aktiviert ist
        if escRate == "YES":
            rente = rente * (1 + (Rate/100))

        # Berechnung der Reserve für nachschüssige Rentenzahlung
        if art == "Nachschüssig":
            barwertfaktor = 0
            barwertfaktornormal = 0
            n = int(n)
            t_alter = int(t_alter)
            for k in range(1, n + 1):
                if t_alter + k >= len(lx):
                    print(f"Error: Attempting to access lx[{t_alter + k}] but len(lx)={len(lx)}")
                    raise IndexError("Index out of bounds while accessing lx.")
                if freq == 1: # Wenn die Häufigkeit der Rentenzahlung einmal jährlich ist
                    barwertfaktor += (v ** k) * (lx[t_alter + k] / lx[t_alter])
                
                else :
                    barwertfaktornormal += (v ** k) * (lx[t_alter + k] / lx[t_alter])
                    
            if freq != 1:
                correction_term = ((((lx[t_alter + n] * (v ** (t_alter + n)))) / (lx[t_alter] * (v ** t_alter))) - 1) * ((freq - 1) / (2 * freq))
                barwertfaktor = barwertfaktornormal - correction_term
            
            barwert = math.ceil(rente * barwertfaktor) # Aufrunden des Barwerts

        
            return barwert
        
        # Berechnung der Reserve für vorschüssige Rentenzahlung
        elif art == "Vorschüssig":
            barwertfaktor = 0
            barwertfaktornormal = 0
            n = int(n)
            t_alter = int(t_alter)
            for k in range(0, n ):
                if t_alter + k >= len(lx):
                    print(f"Error: Attempting to access lx[{t_alter + k}] but len(lx)={len(lx)}")
                    raise IndexError("Index out of bounds while accessing lx.")
                if freq == 1:
                    barwertfaktor += (v ** k) * (lx[t_alter + k] / lx[t_alter])
                
                else :
                    barwertfaktornormal += (v ** k) * (lx[t_alter + k] / lx[t_alter])
                    
            if freq != 1:
                correction_term = ((((lx[t_alter + n] * (v ** (t_alter + n)))) / (lx[t_alter] * (v ** t_alter))) - 1) * ((freq - 1) / (2 * freq))
                barwertfaktor = barwertfaktornormal + correction_term
            
            barwert = math.ceil(rente * barwertfaktor)
        
            return barwert
        

    def compare_reserves(self, directory, base_name='GI_annuities_data_template_with_reserves'):
        file_extension = '.xlsx'
        reserve_sums = {}
 
        for filename in os.listdir(directory):
            if filename.startswith(base_name) and filename.endswith(file_extension):
                file_path = os.path.join(directory, filename)
               
           
                df = pd.read_excel(file_path, sheet_name="Reserves")
               
         
                reserve_sum = df['Reserve'].iloc[-1]
             
                reserve_sums[filename] = reserve_sum
       
        # Wandelt das Dictionary in ein DataFrame um, um den Vergleich zu erleichtern
        comparison_df = pd.DataFrame(list(reserve_sums.items()), columns=['File', 'Sum of Reserves'])
       
        # Sortiert das DataFrame nach dem Dateinamen, um eine richtige Reihenfolge zu gewährleisten
        comparison_df.sort_values(by='File', inplace=True)
       
       # Berechnet die Differenz (Delta) zwischen den Summen der Reserven
        comparison_df['Delta'] = comparison_df['Sum of Reserves'].diff()

        # Berechnet die prozentuale Änderung zwischen den Reserven
        comparison_df['Percentage Change'] = comparison_df['Delta'] / comparison_df['Sum of Reserves'].shift(1) * 100
        
 

        # Speichert den Vergleich in einer neuen Excel-Datei
        comparison_output_path = os.path.join(directory, 'reserves_comparison_with_deltas.xlsx')
        comparison_df.to_excel(comparison_output_path, index=False)
       
        print(f"Comparison of reserves with deltas saved to: {comparison_output_path}")
        return comparison_df
    def plot_reserve_comparison(self, comparison_df, output_directory):
        plt.figure(figsize=(10, 6))
        plt.plot(comparison_df['File'], comparison_df['Sum of Reserves'], marker='o')
        plt.title('Sum of Reserves Comparison')
 
        # Kombiniert die Variablen für die Beschriftung der x-Achse
        xlabel = f"Interest Rate at 0.25% , 0.3% 0.35% reading from left to right, Escalation Rate: {variables_df.at[1, 1]}"
        plt.xlabel(xlabel)
 
        plt.ylabel('Sum of Reserves')
        plt.xticks(rotation=45, ha='right')
        plt.grid(True)
        plt.tight_layout()

        # Speichert das Diagramm als Bild
        output_path = os.path.join(output_directory, 'reserves_comparison_plot.png')
        plt.savefig(output_path)
        plt.show()
 
        logger.info(f"Plot saved to: {output_path}")




# Liest die Excel-Datei ein
file_path = '/Users/alicetangyie/Downloads/Uni/BachelorArbeit/GI_annuities_data_template.v10_GE_2024_Q3_modifiziert Kopie.xlsx'# Muss vom Anwender angepasst werden
bestand_df = pd.read_excel(file_path, sheet_name="Inputs - MPs", header=5)
death_table_df = pd.read_excel(file_path, sheet_name="Death Table", header=3)
 
# Liest den Zinssatz und die Eskalationsrate aus einem anderen Tabellenblatt
variables_df = pd.read_excel(file_path, sheet_name="Variables", header=None)
zins = variables_df.at[0, 1]  # Liest den Wert in Zelle B1  (Zins) zeile-spalte leseweise
Rate = variables_df.at[1, 1]  # Liest den Wert in Zelle B2 (EscRate)
berechnungsjahr = variables_df.at[2, 1]# Liest den Wert in Zelle B3
art = variables_df.at[3, 1]
reserve_berechnung = ReserveBerechnung(death_table=death_table_df)

# Überprüfung der notwendigen Spalten 
required_columns = ['AGE_AT_ENTRY', 'ANN_ANNUITY', 'POL_TERM_Y', 'SEX', 'ESC_RATE', 'ANNUITY_FREQ', 'ENTRY_YEAR',"Q_CORR_PN"]
for col in required_columns:
    if col not in bestand_df.columns:
        raise ValueError(f"Missing required column: {col}")

 # Iteriert über jede Zeile im DataFrame und berechnet die Reserve  
for index, row in bestand_df.iterrows():
    alter = row["AGE_AT_ENTRY"]
    rente = row["ANN_ANNUITY"]
    uebersterblichkeit = row.get("Q_CORR_PN", 1.0) # Verwende den Übersterblichkeitsfaktor oder den Standardwert  
    n = row["POL_TERM_Y"] # Laufzeit der Police
    sex = row["SEX"]
    escRate = row["ESC_RATE"]
    freq = row["ANNUITY_FREQ"]
    geburtsjahr = row["ENTRY_YEAR"]- row["AGE_AT_ENTRY"]

   
    # Berechnet die Reserve
    barwert = reserve_berechnung.berechne_reserve(alter, rente, uebersterblichkeit, zins, n, sex, escRate,freq,geburtsjahr)
    bestand_df.at[index, 'Reserve'] = barwert


 # Summiert die gesamten Reserven
summe_reserven = bestand_df['Reserve'].sum()
 

# Speichert die Ergebnisse in einer neuen Excel-Datei
 
base_name = 'GI_annuities_data_template_with_reserves'
base_directory = '/Users/alicetangyie/Downloads/Uni/BachelorArbeit/'
file_extension = '.xlsx'
file_number = 1

# Überprüft, ob eine Datei bereits existiert, und erhöht die Nummer
while os.path.exists(f"{base_directory}{base_name}_{file_number}{file_extension}"):
    file_number += 1
 
new_file_path = f"{base_directory}{base_name}_{file_number}{file_extension}"

# Speichert das DataFrame in der Excel-Datei
bestand_df.to_excel(new_file_path, sheet_name="Reserves", index=False)

# Öffnet die Arbeitsmappe, um die Summe hinzuzufügen
wb = load_workbook(new_file_path)
ws = wb['Reserves']
 
# Fügt die Summe unterhalb der letzten Zeile der 'Reserve'-Spalte hinzu
last_row = len(bestand_df) + 2  # Hinzufügen von 2, um die Kopfzeile zu berücksichtigen und die 1-basierte Indizierung
sum_cell = ws.cell(row=last_row, column=bestand_df.columns.get_loc('Reserve') + 1)
sum_cell.value = summe_reserven
sum_cell.font = Font(bold=True)  # Text in Fettformat
 
# Speichert die Datei mit der hinzugefügten Summe
wb.save(new_file_path)
 

print(f"Die Ergebnisse wurden in die neue Datei gespeichert: {new_file_path}")
#print(bestand_df.head())
logger.info(f"Process complete: {new_file_path}")    
 
# Verzeichnis festlegen
directory = '/Users/alicetangyie/Downloads/Uni/BachelorArbeit/'

# Vergleich der Reserven und Plotten der Ergebnisse 
reserve_berechnung = ReserveBerechnung(death_table=death_table_df)
comparison_df = reserve_berechnung.compare_reserves(directory)
reserve_berechnung.plot_reserve_comparison(comparison_df, directory)
 
