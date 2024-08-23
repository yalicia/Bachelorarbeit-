import os
import logging
import pandas as pd
import math
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import Font
 
"""
          berechne_reserve  Calculates the reserve for a given policy.
   
            Parameters:
            - alter: Age of the insured
            - rente: Annuity amount
            - uebersterblichkeit: Mortality adjustment factor
            - zins: Interest rate
            - n: Policy term in years
            - sex: Gender (0 for male, 1 for female)
            - escRate: Escalation rate flag ('YES' or 'NO')
            - freq: Annuity frequency
            - geburtsjahr: Year of birth
 
            Returns:
            - barwert: Calculated reserve value
         """  
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
 
class ReserveBerechnung:
    def __init__(self, death_table):
        self.death_table = death_table
    def berechne_reserve(self, alter, rente, uebersterblichkeit, zins, n, sex, escRate,freq,geburtsjahr):
    
        v = 1 / (1 + (zins/100)) # Abzinsfaktor
 
        if sex == 0: # "0" for male gender
            adjusted_qx = self.death_table["q x+0"] * uebersterblichkeit # Übersterblichkeit berücksichtigen
            px = [1 - qx for qx in adjusted_qx]
            lx = [1000000] # Startwert für lx
            for i in range(1, len(px)):
                lx.append(lx[i - 1] * px[i - 1])
            birthyear_row = self.death_table[self.death_table['BIRTH_YEAR'] == geburtsjahr]    
            if not birthyear_row.empty:
                altersverschiebung = birthyear_row['AGE_ADJUSTMENT_M'].values[0]
                alterHeute = berechnungsjahr - geburtsjahr  
                t_alter = alterHeute + altersverschiebung
            else:
                raise IndexError("Der Geburtsjahr für einen Mann wurde falsch gerechnet oder ist nicht zu finden.")
        elif sex == 1: # "1" for female gender
            adjusted_qy = self.death_table["q y+0"] * uebersterblichkeit # Übersterblichkeit berücksichtigen
            py = [1 - qy for qy in adjusted_qy]
            lx = [1000000] # Startwert für lx (using lx to keep consistency)
            for i in range(1, len(py)):
                lx.append(lx[i - 1] * py[i - 1])
            birthyear_row = self.death_table[self.death_table['BIRTH_YEAR'] == geburtsjahr]    
            if not birthyear_row.empty:
                altersverschiebung = birthyear_row['AGE_ADJUSTMENT_F'].values[0]
                alterHeute = berechnungsjahr - geburtsjahr
                t_alter = alterHeute + altersverschiebung
                #print(f"Die Ergebnisse wurden in die natei gespeit: {alterHeute,altersverschiebung,t_alter}")
            else:
                raise IndexError("Der Geburtsjahr für eine Frau wurde falsch gerechnet oder ist nicht zu finden.")
        else:
            return "Ungültige Geschlechtseingabe"
   
        if n == 121:
            n = 121 - t_alter
        #print(f"Die Ergebnisse wurtei gespeit: {n + t_alter}")
        if t_alter + n > len(lx):
            raise IndexError("Der Wert von alter + n überschreitet die Länge von lx.")
        #print(f"Die Ergebnii gespeit: {len(lx)}")
        if escRate == "YES":
            rente = rente * (1 + (Rate/100))
        if art == "Nachschüssig":
            barwertfaktor = 0
            barwertfaktornormal = 0
            n = int(n)
            t_alter = int(t_alter)
            for k in range(1, n + 1):
                if t_alter + k >= len(lx):
                    print(f"Error: Attempting to access lx[{t_alter + k}] but len(lx)={len(lx)}")
                    raise IndexError("Index out of bounds while accessing lx.")
                if freq == 1:
                    barwertfaktor += (v ** k) * (lx[t_alter + k] / lx[t_alter])
                
                else :
                    barwertfaktornormal += (v ** k) * (lx[t_alter + k] / lx[t_alter])
                    #barwertfaktor = barwertfaktornormal - ((((lx[t_alter+n] * (v ** (t_alter+n))))/(lx[t_alter] * (v ** t_alter))-1)* ((freq-1)/2*freq))
            if freq != 1:
                correction_term = ((((lx[t_alter + n] * (v ** (t_alter + n)))) / (lx[t_alter] * (v ** t_alter))) - 1) * ((freq - 1) / (2 * freq))
                barwertfaktor = barwertfaktornormal - correction_term
            
            barwert = math.ceil(rente * barwertfaktor)
        
            return barwert
        elif art == "Vorschüssig":
            barwertfaktor = 0
            barwertfaktornormal = 0
            n = int(n)
            t_alter = int(t_alter)
            for k in range(0, n ):
                if t_alter + k >= len(lx):
                    print(f"Error: Attempting to access lx[{t_alter + k}] but len(lx)={len(lx)}")
                    raise IndexError("Index out of bounds while accessing lx.")
                if freq == 1:
                    barwertfaktor += (v ** k) * (lx[t_alter + k] / lx[t_alter])
                
                else :
                    barwertfaktornormal += (v ** k) * (lx[t_alter + k] / lx[t_alter])
                    #barwertfaktor = barwertfaktornormal - ((((lx[t_alter+n] * (v ** (t_alter+n))))/(lx[t_alter] * (v ** t_alter))-1)* ((freq-1)/2*freq))
            if freq != 1:
                correction_term = ((((lx[t_alter + n] * (v ** (t_alter + n)))) / (lx[t_alter] * (v ** t_alter))) - 1) * ((freq - 1) / (2 * freq))
                barwertfaktor = barwertfaktornormal + correction_term
            
            barwert = math.ceil(rente * barwertfaktor)
        
            return barwert
    def compare_reserves(self, directory, base_name='GI_annuities_data_template_with_reserves'):
        file_extension = '.xlsx'
        reserve_sums = {}
 
        for filename in os.listdir(directory):
            if filename.startswith(base_name) and filename.endswith(file_extension):
                file_path = os.path.join(directory, filename)
               
           
                df = pd.read_excel(file_path, sheet_name="Tabelle2")
               
         
                reserve_sum = df['Reserve'].iloc[-1]
             
                reserve_sums[filename] = reserve_sum
       
        # Convert the dictionary to a DataFrame for easy comparison
        comparison_df = pd.DataFrame(list(reserve_sums.items()), columns=['File', 'Sum of Reserves'])
       
        # Sort the DataFrame by filename to ensure proper order
        comparison_df.sort_values(by='File', inplace=True)
       
       
        comparison_df['Delta'] = comparison_df['Sum of Reserves'].diff()
        comparison_df['Percentage Change'] = comparison_df['Delta'] / comparison_df['Sum of Reserves'].shift(1) * 100
        
 

        # Save the comparison to a new Excel file
        comparison_output_path = os.path.join(directory, 'reserves_comparison_with_deltas.xlsx')
        comparison_df.to_excel(comparison_output_path, index=False)
       
        print(f"Comparison of reserves with deltas saved to: {comparison_output_path}")
        return comparison_df
    def plot_reserve_comparison(self, comparison_df, output_directory):
        plt.figure(figsize=(10, 6))
        plt.plot(comparison_df['File'], comparison_df['Sum of Reserves'], marker='o')
        plt.title('Sum of Reserves Comparison')
 
        # Concatenate the variables for the x-label
        xlabel = f"Interest Rate: {variables_df.at[0, 1]}, Escalation Rate: {variables_df.at[1, 1]}"
        plt.xlabel(xlabel)
 
        plt.ylabel('Sum of Reserves')
        plt.xticks(rotation=45, ha='right')
        plt.grid(True)
        plt.tight_layout()
 
        output_path = os.path.join(output_directory, 'reserves_comparison_plot.png')
        plt.savefig(output_path)
        plt.show()
 
        logger.info(f"Plot saved to: {output_path}")




# Excel-Datei einlesen
file_path = '/Users/alicetangyie/Downloads/Uni/BachelorArbeit/GI_annuities_data_template.v10_GE_2024_Q1.xlsx'# Muss vom Anwender angepasst werden
bestand_df = pd.read_excel(file_path, sheet_name="Inputs - MPs", header=5)
death_table_df = pd.read_excel(file_path, sheet_name="Death Table", header=3)
 
# Lesen des Zinssatzes und der EscRate aus einem anderen Blatt
variables_df = pd.read_excel(file_path, sheet_name="Variables", header=None)
zins = variables_df.at[0, 1]  # Liest den Wert in Zelle B1  (Zins) zeile-spalte leseweise
Rate = variables_df.at[1, 1]  # Liest den Wert in Zelle B2 (EscRate)
berechnungsjahr = variables_df.at[2, 1]# Liest den Wert in Zelle B3
art = variables_df.at[3, 1]
reserve_berechnung = ReserveBerechnung(death_table=death_table_df)
 
required_columns = ['AGE_AT_ENTRY', 'ANN_ANNUITY', 'POL_TERM_Y', 'SEX', 'ESC_RATE', 'ANNUITY_FREQ', 'ENTRY_YEAR',"Q_CORR_PN"]
for col in required_columns:
    if col not in bestand_df.columns:
        raise ValueError(f"Missing required column: {col}")
   
for index, row in bestand_df.iterrows():
    alter = row["AGE_AT_ENTRY"]
    rente = row["ANN_ANNUITY"]
    uebersterblichkeit = row.get("Q_CORR_PN", 1.0) # Verwende den Übersterblichkeitsfaktor oder den Standardwert  
    n = row["POL_TERM_Y"] # Verwende die Laufzeit aus der Tabelle
    sex = row["SEX"]
    escRate = row["ESC_RATE"]
    freq = row["ANNUITY_FREQ"]
    geburtsjahr = row["ENTRY_YEAR"]- row["AGE_AT_ENTRY"]
 

   
 
   
    # Berechnung der Reserve
    barwert = reserve_berechnung.berechne_reserve(alter, rente, uebersterblichkeit, zins, n, sex, escRate,freq,geburtsjahr)
    bestand_df.at[index, 'Reserve'] = barwert
 
summe_reserven = bestand_df['Reserve'].sum()
 

# Speichern der Ergebnisse in einer neuen Excel-Datei
 
base_name = 'GI_annuities_data_template_with_reserves'
base_directory = '/Users/alicetangyie/Downloads/Uni/BachelorArbeit/'
file_extension = '.xlsx'
file_number = 1
 
while os.path.exists(f"{base_directory}{base_name}_{file_number}{file_extension}"):
    file_number += 1
 
new_file_path = f"{base_directory}{base_name}_{file_number}{file_extension}"
 
bestand_df.to_excel(new_file_path, sheet_name="Tabelle2", index=False)
 
wb = load_workbook(new_file_path)
ws = wb['Tabelle2']
 
# Hinzufügen der Summe unterhalb der letzten Zeile der 'Reserve' Spalte
last_row = len(bestand_df) + 2  # Adding 2 to account for the header row and 1-based indexing
sum_cell = ws.cell(row=last_row, column=bestand_df.columns.get_loc('Reserve') + 1)
sum_cell.value = summe_reserven
sum_cell.font = Font(bold=True)  # Text in Fettformat
 
# Speichern der Datei mit der hinzugefügten Summe
wb.save(new_file_path)
 

print(f"Die Ergebnisse wurden in die neue Datei gespeichert: {new_file_path}")
#print(bestand_df.head())
logger.info(f"Process complete: {new_file_path}")    
 

directory = '/Users/alicetangyie/Downloads/Uni/BachelorArbeit/'
 
reserve_berechnung = ReserveBerechnung(death_table=death_table_df)
comparison_df = reserve_berechnung.compare_reserves(directory)
reserve_berechnung.plot_reserve_comparison(comparison_df, directory)
 
# Display the comparison DataFrame
#print(comparison_df)  
'/Users/alicetangyie/Downloads/Uni/BachelorArbeit/GI_annuities_data_template_with_reserves.xlsx'